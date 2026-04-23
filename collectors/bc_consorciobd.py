#!/usr/bin/env python3
"""
Coletor do ConsorcioBD (BCB) por descoberta de links na página oficial.

Esta versão foi endurecida para diagnóstico:
- salva snapshot do HTML recebido;
- tenta descobrir candidatos por <a href> e por regex no HTML bruto;
- grava runtime mesmo quando falha;
- expõe evidências mínimas para depuração sem chute.

Este coletor não gera data/dist. Isso fica para o build-read-models.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import sys
import zipfile
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from unidecode import unidecode
from urllib3.util.retry import Retry


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def dump_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2, sort_keys=False)
        fh.write("\n")


def append_github_output(name: str, value: str) -> None:
    github_output = os.environ.get("GITHUB_OUTPUT")
    if not github_output:
        return
    with open(github_output, "a", encoding="utf-8") as fh:
        fh.write(f"{name}={value}\n")


def stable_json_dumps(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_of_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_of(data: Any) -> str:
    return hashlib.sha256(stable_json_dumps(data).encode("utf-8")).hexdigest()


def normalize_spaces(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def normalized_match_text(value: str) -> str:
    text = unidecode(str(value or "")).lower()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def get_source_config(config: Dict[str, Any], source_name: str) -> Dict[str, Any]:
    try:
        return config["sources"][source_name]
    except KeyError as exc:
        raise KeyError(f"Fonte '{source_name}' não encontrada em config/sources.json.") from exc


def build_session(timeout_seconds: int) -> requests.Session:
    retry = Retry(
        total=3,
        connect=3,
        read=3,
        backoff_factor=1.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET", "HEAD"}),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)

    session = requests.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.request_timeout = timeout_seconds  # type: ignore[attr-defined]
    return session


def guess_extension_from_url(url: str, accept_extensions: List[str]) -> Optional[str]:
    path = urlparse(url).path.lower()
    for ext in accept_extensions:
        if path.endswith(ext.lower()):
            return ext.lower()
    return None


def infer_temporal_rank(text: str) -> Tuple[int, int, int]:
    text_norm = str(text)

    patterns = [
        r"(20\d{2})[-_/](\d{1,2})[-_/](\d{1,2})",
        r"(\d{1,2})[-_/](\d{1,2})[-_/](20\d{2})",
        r"(20\d{2})(\d{2})(\d{2})",
        r"(20\d{2})[-_/](\d{1,2})",
        r"(20\d{2})(\d{2})",
    ]

    for pattern in patterns:
        match = re.search(pattern, text_norm)
        if not match:
            continue

        parts = match.groups()
        if len(parts) == 3:
            a, b, c = parts
            if len(a) == 4 and a.startswith("20"):
                year, month, day = int(a), int(b), int(c)
            elif len(c) == 4 and c.startswith("20"):
                day, month, year = int(a), int(b), int(c)
            else:
                continue
            month = max(1, min(month, 12))
            day = max(1, min(day, 31))
            return year, month, day

        if len(parts) == 2:
            year, month = int(parts[0]), int(parts[1])
            month = max(1, min(month, 12))
            return year, month, 1

    year_match = re.search(r"(20\d{2})", text_norm)
    if year_match:
        return int(year_match.group(1)), 1, 1

    return 0, 0, 0


def build_candidate(
    *,
    index: int,
    href: str,
    resolved_url: str,
    text_blob: str,
    accept_extensions: List[str],
    match_any: List[str],
    source_kind: str,
) -> Optional[Dict[str, Any]]:
    ext = guess_extension_from_url(resolved_url, accept_extensions)
    if not ext:
        return None

    text_norm = normalized_match_text(text_blob)
    match_terms = [normalized_match_text(term) for term in match_any]
    matched_terms = [term for term in match_terms if term and term in text_norm]
    if not matched_terms:
        return None

    score = 0
    score += 100 * len(matched_terms)

    if "consorciobd" in text_norm:
        score += 50

    if ext == ".zip":
        score += 20
    elif ext == ".xlsx":
        score += 10
    elif ext == ".csv":
        score += 5

    temporal_rank = infer_temporal_rank(text_blob)

    return {
        "index": index,
        "source_kind": source_kind,
        "href": href,
        "resolved_url": resolved_url,
        "text_blob": normalize_spaces(text_blob) or "",
        "extension": ext,
        "matched_terms": matched_terms,
        "score": score,
        "temporal_rank": list(temporal_rank),
    }


def discover_candidates_from_anchors(
    html: str,
    page_url: str,
    accept_extensions: List[str],
    match_any: List[str],
) -> List[Dict[str, Any]]:
    soup = BeautifulSoup(html, "lxml")
    candidates: List[Dict[str, Any]] = []

    for index, anchor in enumerate(soup.find_all("a", href=True), start=1):
        href = normalize_spaces(anchor.get("href"))
        if not href:
            continue
        resolved_url = urljoin(page_url, href)
        link_text = normalize_spaces(anchor.get_text(" ", strip=True)) or ""
        text_blob = f"{resolved_url} {link_text}"

        candidate = build_candidate(
            index=index,
            href=href,
            resolved_url=resolved_url,
            text_blob=text_blob,
            accept_extensions=accept_extensions,
            match_any=match_any,
            source_kind="anchor",
        )
        if candidate:
            candidates.append(candidate)

    return candidates


def discover_candidates_from_raw_html(
    html: str,
    page_url: str,
    accept_extensions: List[str],
    match_any: List[str],
) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []
    seen_urls: set[str] = set()

    url_pattern = re.compile(
        r"""(?P<url>
            https?://[^\s"'<>]+
            |
            /[^\s"'<>]+
        )""",
        re.IGNORECASE | re.VERBOSE,
    )

    for index, match in enumerate(url_pattern.finditer(html), start=1):
        raw_url = match.group("url")
        if not raw_url:
            continue

        resolved_url = urljoin(page_url, raw_url)
        if resolved_url in seen_urls:
            continue
        seen_urls.add(resolved_url)

        start = max(0, match.start() - 250)
        end = min(len(html), match.end() + 250)
        context = html[start:end]
        text_blob = f"{resolved_url} {context}"

        candidate = build_candidate(
            index=index,
            href=raw_url,
            resolved_url=resolved_url,
            text_blob=text_blob,
            accept_extensions=accept_extensions,
            match_any=match_any,
            source_kind="raw_html",
        )
        if candidate:
            candidates.append(candidate)

    return candidates


def dedupe_and_sort_candidates(candidates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: Dict[str, Dict[str, Any]] = {}

    for item in candidates:
        key = item["resolved_url"]
        current = deduped.get(key)
        if current is None:
            deduped[key] = item
            continue

        current_rank = (tuple(current["temporal_rank"]), current["score"], -current["index"])
        new_rank = (tuple(item["temporal_rank"]), item["score"], -item["index"])
        if new_rank > current_rank:
            deduped[key] = item

    final = list(deduped.values())
    final.sort(
        key=lambda item: (
            tuple(item["temporal_rank"]),
            item["score"],
            -item["index"],
        ),
        reverse=True,
    )
    return final


def select_candidate(candidates: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not candidates:
        raise ValueError("Nenhum candidato compatível encontrado na página oficial do ConsorcioBD.")
    return candidates[0]


def detect_csv_delimiter(text: str) -> str:
    sample = text[:10000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
        return dialect.delimiter
    except Exception:
        if sample.count(";") > sample.count(","):
            return ";"
        return ","


def inspect_binary_payload(binary: bytes, content_type: Optional[str]) -> Dict[str, Any]:
    info: Dict[str, Any] = {
        "detected_format": "binary",
        "content_type": content_type,
        "size_bytes": len(binary),
    }

    if binary.startswith(b"PK\x03\x04"):
        try:
            with zipfile.ZipFile(io.BytesIO(binary)) as zf:
                members = zf.namelist()
                info["archive_members"] = members

                if any(name.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")) for name in members):
                    info["detected_format"] = "zip_with_excel"
                elif any(name.lower().endswith(".csv") for name in members):
                    info["detected_format"] = "zip_with_csv"
                else:
                    try:
                        wb = load_workbook(io.BytesIO(binary), read_only=True, data_only=True)
                        info["detected_format"] = "xlsx"
                        info["workbook_sheets"] = list(wb.sheetnames)
                        wb.close()
                        return info
                    except Exception:
                        info["detected_format"] = "zip"
                return info
        except Exception:
            pass

        try:
            wb = load_workbook(io.BytesIO(binary), read_only=True, data_only=True)
            info["detected_format"] = "xlsx"
            info["workbook_sheets"] = list(wb.sheetnames)
            wb.close()
            return info
        except Exception:
            pass

    try:
        for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                text = binary.decode(encoding)
                delimiter = detect_csv_delimiter(text)
                reader = csv.reader(io.StringIO(text), delimiter=delimiter)
                header = next(reader, [])
                if header:
                    info["detected_format"] = "csv"
                    info["csv_encoding"] = encoding
                    info["csv_delimiter"] = delimiter
                    info["csv_header"] = header[:50]
                    return info
            except Exception:
                continue
    except Exception:
        pass

    return info


def read_existing_source_sha(path: Path) -> Optional[str]:
    if not path.exists():
        return None
    return sha256_of_bytes(path.read_bytes())


def preview_text(text: str, limit: int = 500) -> str:
    compact = " ".join((text or "").split())
    return compact[:limit]


def main() -> int:
    parser = argparse.ArgumentParser(description="Coletor do ConsorcioBD por descoberta HTML.")
    parser.add_argument("--config", required=True, help="Caminho para config/sources.json")
    parser.add_argument("--source", default="bc_consorciobd", help="Nome da fonte em config/sources.json")
    args = parser.parse_args()

    config_path = Path(args.config)
    config = load_json(config_path)
    source_cfg = get_source_config(config, args.source)
    defaults = config.get("defaults", {})

    discovery_cfg = source_cfg["discovery"]
    storage_cfg = source_cfg["storage"]

    raw_dir = Path(storage_cfg["raw_dir"])
    stage_dir = Path(storage_cfg["stage_dir"])
    runtime_dir = Path("data/runtime")

    raw_file = raw_dir / "latest.json"
    raw_binary_file = raw_dir / "latest_source.bin"
    stage_file = stage_dir / "consorciobd_catalog.json"
    runtime_file = runtime_dir / "bc_consorciobd.json"
    snapshot_file = raw_dir / "page_snapshot.html"

    raw_dir.mkdir(parents=True, exist_ok=True)
    stage_dir.mkdir(parents=True, exist_ok=True)
    runtime_dir.mkdir(parents=True, exist_ok=True)

    if not source_cfg.get("enabled", False):
        dump_json(
            runtime_file,
            {
                "source": args.source,
                "last_checked_at": utc_now_iso(),
                "changed": False,
                "mode_used": "disabled",
                "raw_file": str(raw_file),
                "raw_binary_file": str(raw_binary_file),
                "stage_file": str(stage_file),
                "snapshot_file": str(snapshot_file),
            },
        )
        append_github_output("changed", "false")
        append_github_output("records", "0")
        append_github_output("mode_used", "disabled")
        print(f"Fonte '{args.source}' está desabilitada.")
        return 0

    page_url = discovery_cfg["page_url"]
    accept_extensions = discovery_cfg.get("accept_extensions", [".zip", ".csv", ".xlsx"])
    match_any = discovery_cfg.get("match_any", ["ConsorcioBD", "consorciobd"])

    user_agent = defaults.get(
        "user_agent",
        "comparador-consorcios-data/1.0 (+https://sanida.com.br/financas/consorcio/)",
    )
    timeout_seconds = int(defaults.get("timeout_seconds", 60))

    headers_html = {
        "Accept": "text/html,application/xhtml+xml",
        "User-Agent": user_agent,
    }
    headers_file = {
        "Accept": "*/*",
        "User-Agent": user_agent,
    }

    collected_at = utc_now_iso()
    session = build_session(timeout_seconds=timeout_seconds)

    mode_used = "html-link-discovery-failed"
    changed = False
    candidate_count = 0

    try:
        page_response = session.get(
            page_url,
            headers=headers_html,
            timeout=getattr(session, "request_timeout", 60),
        )
        page_response.raise_for_status()

        html_text = page_response.text or ""
        snapshot_file.write_text(html_text, encoding="utf-8")

        anchor_candidates = discover_candidates_from_anchors(
            html=html_text,
            page_url=page_url,
            accept_extensions=accept_extensions,
            match_any=match_any,
        )
        raw_html_candidates = discover_candidates_from_raw_html(
            html=html_text,
            page_url=page_url,
            accept_extensions=accept_extensions,
            match_any=match_any,
        )
        candidates = dedupe_and_sort_candidates(anchor_candidates + raw_html_candidates)
        candidate_count = len(candidates)

        if not candidates:
            runtime_payload = {
                "source": args.source,
                "last_checked_at": collected_at,
                "changed": False,
                "mode_used": mode_used,
                "page_url": page_url,
                "candidate_count": 0,
                "html_preview": preview_text(html_text),
                "snapshot_file": str(snapshot_file),
                "raw_file": str(raw_file),
                "raw_binary_file": str(raw_binary_file),
                "stage_file": str(stage_file),
            }
            dump_json(runtime_file, runtime_payload)
            append_github_output("changed", "false")
            append_github_output("records", "0")
            append_github_output("mode_used", mode_used)
            raise ValueError(
                "Nenhum candidato compatível encontrado na página oficial do ConsorcioBD. "
                f"Snapshot salvo em {snapshot_file}."
            )

        selected = select_candidate(candidates)

        download_response = session.get(
            selected["resolved_url"],
            headers=headers_file,
            timeout=getattr(session, "request_timeout", 120),
            allow_redirects=True,
        )
        download_response.raise_for_status()

        binary = download_response.content
        if not binary:
            raise ValueError("O recurso selecionado do ConsorcioBD retornou conteúdo vazio.")

        source_sha = sha256_of_bytes(binary)
        previous_source_sha = read_existing_source_sha(raw_binary_file)
        changed = previous_source_sha != source_sha
        mode_used = "html-link-download"

        inspection = inspect_binary_payload(binary, download_response.headers.get("Content-Type"))

        selected_resource = {
            "page_url": page_url,
            "selected_candidate": selected,
            "final_url": download_response.url,
            "original_content_type": download_response.headers.get("Content-Type"),
            "original_content_length": download_response.headers.get("Content-Length"),
            "sha256": source_sha,
            "inspection": inspection,
        }

        raw_payload = {
            "metadata": {
                "source": args.source,
                "collector": "collectors/bc_consorciobd.py",
                "collected_at": collected_at,
                "mode_used": mode_used,
                "page_url": page_url,
                "candidate_count": len(candidates),
                "resource_sha256": source_sha,
            },
            "selected_resource": selected_resource,
            "candidates": candidates,
        }

        stage_payload = {
            "metadata": {
                "source": args.source,
                "collector": "collectors/bc_consorciobd.py",
                "collected_at": collected_at,
                "mode_used": mode_used,
                "page_url": page_url,
                "candidate_count": len(candidates),
                "resource_sha256": source_sha,
            },
            "resource": selected_resource,
            "candidates": candidates,
        }

        runtime_payload = {
            "source": args.source,
            "last_checked_at": collected_at,
            "last_changed_at": collected_at if changed else None,
            "changed": changed,
            "mode_used": mode_used,
            "page_url": page_url,
            "candidate_count": len(candidates),
            "resource_sha256": source_sha,
            "selected_url": selected["resolved_url"],
            "final_url": download_response.url,
            "snapshot_file": str(snapshot_file),
            "raw_file": str(raw_file),
            "raw_binary_file": str(raw_binary_file),
            "stage_file": str(stage_file),
        }

        if changed:
            raw_binary_file.write_bytes(binary)
            dump_json(raw_file, raw_payload)
            dump_json(stage_file, stage_payload)

        dump_json(runtime_file, runtime_payload)

        append_github_output("changed", "true" if changed else "false")
        append_github_output("records", str(len(candidates)))
        append_github_output("stage_hash", sha256_of(stage_payload))
        append_github_output("mode_used", mode_used)

        print(
            json.dumps(
                {
                    "source": args.source,
                    "changed": changed,
                    "records": len(candidates),
                    "mode_used": mode_used,
                    "raw_file": str(raw_file),
                    "raw_binary_file": str(raw_binary_file),
                    "stage_file": str(stage_file),
                    "runtime_file": str(runtime_file),
                    "snapshot_file": str(snapshot_file),
                },
                ensure_ascii=False,
            )
        )
        return 0

    except requests.HTTPError as exc:
        dump_json(
            runtime_file,
            {
                "source": args.source,
                "last_checked_at": collected_at,
                "changed": False,
                "mode_used": mode_used,
                "page_url": page_url,
                "candidate_count": candidate_count,
                "error": f"Erro HTTP ao coletar ConsorcioBD: {exc}",
                "snapshot_file": str(snapshot_file),
                "raw_file": str(raw_file),
                "raw_binary_file": str(raw_binary_file),
                "stage_file": str(stage_file),
            },
        )
        append_github_output("changed", "false")
        append_github_output("records", "0")
        append_github_output("mode_used", mode_used)
        print(f"Erro HTTP ao coletar ConsorcioBD: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        dump_json(
            runtime_file,
            {
                "source": args.source,
                "last_checked_at": collected_at,
                "changed": False,
                "mode_used": mode_used,
                "page_url": page_url,
                "candidate_count": candidate_count,
                "error": str(exc),
                "snapshot_file": str(snapshot_file),
                "raw_file": str(raw_file),
                "raw_binary_file": str(raw_binary_file),
                "stage_file": str(stage_file),
            },
        )
        append_github_output("changed", "false")
        append_github_output("records", "0")
        append_github_output("mode_used", mode_used)
        print(f"Falha no coletor bc_consorciobd: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
