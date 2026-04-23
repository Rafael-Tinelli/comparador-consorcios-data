#!/usr/bin/env python3
"""
Coletor do ConsorcioBD trimestral via URL direta confirmada.

Escopo desta versão:
- coleta os dois fluxos trimestrais do ConsorcioBD:
  - Dados por Unidade da Federação (UF)
  - Dados Contábeis Consolidados (ADM)
- usa probing determinístico sobre /Fis/Consorcios/Port/BD/;
- grava raw, stage e runtime;
- só considera sucesso quando mode_used = "direct-content-download".
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import sys
import zipfile
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def dump_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2, sort_keys=False)
        fh.write("\n")


def stable_json_dumps(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_of(data: Any) -> str:
    return hashlib.sha256(stable_json_dumps(data).encode("utf-8")).hexdigest()


def sha256_of_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def append_github_output(name: str, value: str) -> None:
    github_output = os.environ.get("GITHUB_OUTPUT")
    if not github_output:
        return
    with open(github_output, "a", encoding="utf-8") as fh:
        fh.write(f"{name}={value}\n")


def get_source_config(config: Dict[str, Any], source_name: str) -> Dict[str, Any]:
    try:
        return config["sources"][source_name]
    except KeyError as exc:
        raise KeyError(f"Fonte '{source_name}' não encontrada em config/sources.json.") from exc


def normalize_spaces(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def build_session(timeout_seconds: int) -> requests.Session:
    retry = Retry(
        total=3,
        connect=3,
        read=3,
        backoff_factor=1.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET", "HEAD"}),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)

    session = requests.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.request_timeout = timeout_seconds  # type: ignore[attr-defined]
    return session


def iter_year_months(reference: date, months_back: int) -> List[Tuple[int, int]]:
    items: List[Tuple[int, int]] = []
    year = reference.year
    month = reference.month

    for _ in range(months_back):
        items.append((year, month))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    return items


def render_url(base_url: str, filename_pattern: str, year: int, month: int) -> str:
    yyyymm = f"{year}{month:02d}"
    filename = filename_pattern.format(
        yyyy=f"{year}",
        mm=f"{month:02d}",
        yyyymm=yyyymm,
    )
    return f"{base_url.rstrip('/')}/{filename}"


def candidate_looks_valid_from_headers(
    *,
    status_code: Optional[int],
    final_url: str,
    content_type: Optional[str],
    content_length: Optional[str],
    min_size_bytes: int,
) -> bool:
    if status_code != 200:
        return False

    final_url_l = final_url.lower()
    if not final_url_l.endswith(".zip"):
        return False

    ct = (content_type or "").lower()
    if "text/html" in ct:
        return False

    if content_length:
        try:
            if int(content_length) < min_size_bytes:
                return False
        except Exception:
            pass

    return True


def probe_head(
    session: requests.Session,
    url: str,
    headers: Dict[str, str],
    min_size_bytes: int,
) -> Dict[str, Any]:
    try:
        response = session.head(
            url,
            headers=headers,
            timeout=getattr(session, "request_timeout", 60),
            allow_redirects=True,
        )
        result = {
            "method": "HEAD",
            "status_code": response.status_code,
            "final_url": response.url,
            "content_type": response.headers.get("Content-Type"),
            "content_length": response.headers.get("Content-Length"),
        }
        result["ok"] = candidate_looks_valid_from_headers(
            status_code=result["status_code"],
            final_url=result["final_url"],
            content_type=result["content_type"],
            content_length=result["content_length"],
            min_size_bytes=min_size_bytes,
        )
        return result
    except Exception as exc:
        return {
            "method": "HEAD",
            "status_code": None,
            "final_url": url,
            "content_type": None,
            "content_length": None,
            "ok": False,
            "error": str(exc),
        }


def probe_get(
    session: requests.Session,
    url: str,
    headers: Dict[str, str],
    min_size_bytes: int,
) -> Dict[str, Any]:
    try:
        response = session.get(
            url,
            headers=headers,
            timeout=getattr(session, "request_timeout", 90),
            allow_redirects=True,
            stream=True,
        )

        first_chunk = b""
        try:
            for chunk in response.iter_content(chunk_size=4096):
                if chunk:
                    first_chunk = chunk
                    break
        finally:
            response.close()

        result = {
            "method": "GET",
            "status_code": response.status_code,
            "final_url": response.url,
            "content_type": response.headers.get("Content-Type"),
            "content_length": response.headers.get("Content-Length"),
            "first_chunk_sha256": sha256_of_bytes(first_chunk) if first_chunk else None,
            "first_chunk_preview_hex": first_chunk[:32].hex() if first_chunk else None,
        }
        result["ok"] = candidate_looks_valid_from_headers(
            status_code=result["status_code"],
            final_url=result["final_url"],
            content_type=result["content_type"],
            content_length=result["content_length"],
            min_size_bytes=min_size_bytes,
        )
        return result
    except Exception as exc:
        return {
            "method": "GET",
            "status_code": None,
            "final_url": url,
            "content_type": None,
            "content_length": None,
            "ok": False,
            "error": str(exc),
        }


def detect_csv_delimiter(text: str) -> str:
    sample = text[:10000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
        return dialect.delimiter
    except Exception:
        return ";" if sample.count(";") > sample.count(",") else ","


def inspect_binary_payload(binary: bytes, content_type: Optional[str]) -> Dict[str, Any]:
    info: Dict[str, Any] = {
        "detected_format": "binary",
        "content_type": content_type,
        "size_bytes": len(binary),
    }

    if binary.startswith(b"PK\x03\x04"):
        try:
            with zipfile.ZipFile(io.BytesIO(binary)) as zf:
                members = zf.namelist()
                info["archive_members"] = members

                lower_members = [name.lower() for name in members]
                if any(name.endswith(".csv") for name in lower_members):
                    info["detected_format"] = "zip_with_csv"
                elif any(name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")) for name in lower_members):
                    info["detected_format"] = "zip_with_excel"
                else:
                    info["detected_format"] = "zip"

                info["archive_members_preview"] = members[:8]
                return info
        except Exception:
            pass

        try:
            wb = load_workbook(io.BytesIO(binary), read_only=True, data_only=True)
            info["detected_format"] = "xlsx"
            info["workbook_sheets"] = list(wb.sheetnames)
            wb.close()
            return info
        except Exception:
            pass

    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = binary.decode(encoding)
            delimiter = detect_csv_delimiter(text)
            reader = csv.reader(io.StringIO(text), delimiter=delimiter)
            header = next(reader, [])
            if header:
                info["detected_format"] = "csv"
                info["csv_encoding"] = encoding
                info["csv_delimiter"] = delimiter
                info["csv_header"] = header[:50]
                return info
        except Exception:
            continue

    return info


def download_binary(session: requests.Session, url: str, headers: Dict[str, str]) -> Tuple[bytes, Dict[str, Any]]:
    response = session.get(
        url,
        headers=headers,
        timeout=getattr(session, "request_timeout", 180),
        allow_redirects=True,
    )
    response.raise_for_status()

    binary = response.content
    if not binary:
        raise ValueError("O download final retornou conteúdo vazio.")

    meta = {
        "status_code": response.status_code,
        "final_url": response.url,
        "content_type": response.headers.get("Content-Type"),
        "content_length": response.headers.get("Content-Length"),
        "last_modified": response.headers.get("Last-Modified"),
        "etag": response.headers.get("ETag"),
    }
    return binary, meta


def read_existing_source_sha(path: Path) -> Optional[str]:
    if not path.exists():
        return None
    return sha256_of_bytes(path.read_bytes())


def find_latest_family_resource(
    *,
    session: requests.Session,
    base_url: str,
    family: Dict[str, Any],
    headers: Dict[str, str],
    reference_date: date,
    http_methods: List[str],
) -> Dict[str, Any]:
    family_key = family["key"]
    family_label = family["label"]
    filename_pattern = family["filename_pattern"]
    allowed_months = set(int(m) for m in family.get("allowed_months", [3, 6, 9, 12]))
    probe_months_back = int(family.get("probe_months_back", 24))
    min_size_bytes = int(family.get("min_size_bytes", 1024))

    probe_results: List[Dict[str, Any]] = []
    selected_candidate: Optional[Dict[str, Any]] = None
    selected_probe: Optional[Dict[str, Any]] = None

    for year, month in iter_year_months(reference_date, probe_months_back):
        if month not in allowed_months:
            continue

        yyyymm = f"{year}{month:02d}"
        url = render_url(base_url, filename_pattern, year, month)
        candidate = {
            "family_key": family_key,
            "family_label": family_label,
            "year": year,
            "month": month,
            "yyyymm": yyyymm,
            "url": url,
        }

        probes: List[Dict[str, Any]] = []
        head_result: Optional[Dict[str, Any]] = None
        get_result: Optional[Dict[str, Any]] = None

        if "HEAD" in http_methods:
            head_result = probe_head(session, url, headers, min_size_bytes)
            probes.append(head_result)

        if head_result and head_result.get("ok"):
            selected_candidate = candidate
            selected_probe = head_result
            probe_results.append({"candidate": candidate, "probes": probes})
            break

        if "GET" in http_methods:
            get_result = probe_get(session, url, headers, min_size_bytes)
            probes.append(get_result)

        probe_results.append({"candidate": candidate, "probes": probes})

        if get_result and get_result.get("ok"):
            selected_candidate = candidate
            selected_probe = get_result
            break

    if not selected_candidate:
        raise ValueError(
            f"Nenhuma URL trimestral homologável respondeu como ZIP válido para a família '{family_key}'."
        )

    binary, download_meta = download_binary(session, selected_candidate["url"], headers)
    inspection = inspect_binary_payload(binary, download_meta.get("content_type"))

    return {
        "family_key": family_key,
        "family_label": family_label,
        "probe_results": probe_results,
        "selected_candidate": selected_candidate,
        "selected_probe": selected_probe,
        "binary": binary,
        "download_meta": download_meta,
        "inspection": inspection,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Coletor do ConsorcioBD trimestral via URL direta.")
    parser.add_argument("--config", required=True, help="Caminho para config/sources.json")
    parser.add_argument("--source", default="bc_consorciobd_trimestral", help="Nome da fonte em config/sources.json")
    args = parser.parse_args()

    config = load_json(Path(args.config))
    source_cfg = get_source_config(config, args.source)
    defaults = config.get("defaults", {})

    raw_dir = Path(source_cfg["storage"]["raw_dir"])
    stage_dir = Path(source_cfg["storage"]["stage_dir"])
    runtime_dir = Path("data/runtime")

    raw_file = raw_dir / "latest.json"
    stage_file = stage_dir / "consorciobd_trimestral_catalog.json"
    runtime_file = runtime_dir / "bc_consorciobd_trimestral.json"

    raw_dir.mkdir(parents=True, exist_ok=True)
    stage_dir.mkdir(parents=True, exist_ok=True)
    runtime_dir.mkdir(parents=True, exist_ok=True)

    collected_at = utc_now_iso()
    mode_used = "direct-probe-failed"

    if not source_cfg.get("enabled", False):
        runtime_payload = {
            "source": args.source,
            "last_checked_at": collected_at,
            "changed": False,
            "mode_used": "disabled",
            "raw_file": str(raw_file),
            "stage_file": str(stage_file),
        }
        dump_json(runtime_file, runtime_payload)
        append_github_output("changed", "false")
        append_github_output("records", "0")
        append_github_output("mode_used", "disabled")
        print(f"Fonte '{args.source}' está desabilitada.")
        return 0

    direct_cfg = source_cfg["direct_download"]
    families = [f for f in direct_cfg.get("families", []) if f.get("enabled", True)]
    if not families:
        raise ValueError("Nenhuma família ativa encontrada em direct_download.families.")

    user_agent = defaults.get(
        "user_agent",
        "comparador-consorcios-data/1.0 (+https://sanida.com.br/financas/consorcio/)",
    )
    timeout_seconds = int(defaults.get("timeout_seconds", 60))
    session = build_session(timeout_seconds=timeout_seconds)

    headers = {
        "Accept": "*/*",
        "User-Agent": user_agent,
        "Referer": source_cfg["official_page_url"],
    }

    base_url = direct_cfg["base_url"]
    http_methods = [m.upper() for m in direct_cfg.get("http_methods", ["HEAD", "GET"])]
    reference_date = datetime.now(timezone.utc).date()

    family_results: List[Dict[str, Any]] = []
    all_probe_results: Dict[str, List[Dict[str, Any]]] = {}
    changed_any = False

    try:
        for family in families:
            result = find_latest_family_resource(
                session=session,
                base_url=base_url,
                family=family,
                headers=headers,
                reference_date=reference_date,
                http_methods=http_methods,
            )

            binary = result["binary"]
            family_key = result["family_key"]
            binary_file = raw_dir / f"latest_{family_key}.bin"
            source_sha = sha256_of_bytes(binary)
            previous_source_sha = read_existing_source_sha(binary_file)
            family_changed = previous_source_sha != source_sha

            if family_changed:
                binary_file.write_bytes(binary)

            changed_any = changed_any or family_changed
            all_probe_results[family_key] = result["probe_results"]

            family_results.append(
                {
                    "family_key": family_key,
                    "family_label": result["family_label"],
                    "selected_candidate": result["selected_candidate"],
                    "selected_probe": result["selected_probe"],
                    "download_meta": result["download_meta"],
                    "inspection": result["inspection"],
                    "sha256": source_sha,
                    "binary_file": str(binary_file),
                    "changed": family_changed,
                }
            )

        mode_used = "direct-content-download"

        raw_payload = {
            "metadata": {
                "source": args.source,
                "collector": "collectors/bc_consorciobd_trimestral.py",
                "collected_at": collected_at,
                "mode_used": mode_used,
                "family_count": len(family_results)
            },
            "resources": family_results,
            "probe_results_by_family": all_probe_results,
        }

        stage_payload = {
            "metadata": {
                "source": args.source,
                "collector": "collectors/bc_consorciobd_trimestral.py",
                "collected_at": collected_at,
                "mode_used": mode_used,
                "family_count": len(family_results)
            },
            "resources": family_results,
            "probe_results_by_family": all_probe_results,
        }

        runtime_payload = {
            "source": args.source,
            "last_checked_at": collected_at,
            "last_changed_at": collected_at if changed_any else None,
            "changed": changed_any,
            "mode_used": mode_used,
            "family_count": len(family_results),
            "resources": [
                {
                    "family_key": item["family_key"],
                    "family_label": item["family_label"],
                    "selected_url": item["selected_candidate"]["url"],
                    "final_url": item["download_meta"].get("final_url"),
                    "sha256": item["sha256"],
                    "changed": item["changed"],
                    "binary_file": item["binary_file"],
                }
                for item in family_results
            ],
            "raw_file": str(raw_file),
            "stage_file": str(stage_file),
        }

        if changed_any:
            dump_json(raw_file, raw_payload)
            dump_json(stage_file, stage_payload)

        dump_json(runtime_file, runtime_payload)

        append_github_output("changed", "true" if changed_any else "false")
        append_github_output("records", str(len(family_results)))
        append_github_output("stage_hash", sha256_of(stage_payload))
        append_github_output("mode_used", mode_used)

        print(
            json.dumps(
                {
                    "source": args.source,
                    "changed": changed_any,
                    "records": len(family_results),
                    "mode_used": mode_used,
                    "raw_file": str(raw_file),
                    "stage_file": str(stage_file),
                    "runtime_file": str(runtime_file),
                },
                ensure_ascii=False,
            )
        )
        return 0

    except requests.HTTPError as exc:
        runtime_payload = {
            "source": args.source,
            "last_checked_at": collected_at,
            "changed": False,
            "mode_used": mode_used,
            "error": f"Erro HTTP ao coletar ConsorcioBD trimestral: {exc}",
            "raw_file": str(raw_file),
            "stage_file": str(stage_file),
        }
        dump_json(runtime_file, runtime_payload)
        append_github_output("changed", "false")
        append_github_output("records", "0")
        append_github_output("mode_used", mode_used)
        print(f"Erro HTTP ao coletar ConsorcioBD trimestral: {exc}", file=sys.stderr)
        return 1

    except Exception as exc:
        runtime_payload = {
            "source": args.source,
            "last_checked_at": collected_at,
            "changed": False,
            "mode_used": mode_used,
            "error": str(exc),
            "raw_file": str(raw_file),
            "stage_file": str(stage_file),
        }
        dump_json(runtime_file, runtime_payload)
        append_github_output("changed", "false")
        append_github_output("records", "0")
        append_github_output("mode_used", mode_used)
        print(f"Falha no coletor bc_consorciobd_trimestral: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
